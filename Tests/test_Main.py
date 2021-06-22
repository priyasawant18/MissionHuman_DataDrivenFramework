from time import sleep

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

from Utils.Utilities import Testlinks
from Utils.Utilities import Intantiate
import openpyxl

Dict = {}
Dict_state = {}
Dict_district = {}
Dict_resourse = {}
Dict_data = {}

def test_setup():
    Intantiate.driver.get(Testlinks.BASE_URL)
    sleep(2)

def test_extract_state():
    state_name = Intantiate.driver.find_elements_by_xpath("//div[@class='location-column']")
    print(len(state_name))
    for val in state_name:
       val.find_element_by_xpath("//div[@class='location-column']")
       Dict_state["State"] = val.text
       #print(Dict1)
       val.click()
       sleep(2)
       test_extract_district()


def test_loadExcel():
    book = openpyxl.load_workbook("C:\\Training matrial\\Project Covid\\MIssionHumanData.xlsx")
    sheet = book.active
    cell = sheet.cell(row=2, column=1)
    for i in range(1, sheet.max_row + 1):  # to get rows
        if sheet.cell(row=i, column=2).value == "Andhra Pradesh":
            for j in range(1, sheet.max_column + 1):  # to get column
                # print(sheet.cell(row=i, column=1).value)
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    print(Dict)

def test_StateCompare():
    for key in list(Dict_state.keys()):
        if key in list(Dict.keys()):
            print(True)
        else:
            print(False)

def test_extract_district():
    district_name = Intantiate.driver.find_elements_by_xpath("//div[@class='location-column']")
    print(len(district_name))
    for district in district_name:
        district.find_element_by_xpath("//div[@class='location-column']")
        Dict_district["District"] = district.text
        print(Dict_district)
        district.click()
        test_extract_resourse()
    #Intantiate.driver.find_element_by_xpath("//div[@class='bookmarked' and contains(text(), 'MissionHumane.org')]").click()
    sleep(2)

def test_extract_resourse():
    resourse_name = Intantiate.driver.find_elements_by_xpath("//div[@class='sc-hKFxyN kksiKu']")
    print(len(resourse_name))
    for resourse in resourse_name:
        resourse.find_element_by_xpath("//div[@class='sc-hKFxyN kksiKu']")
        Dict_resourse["Resourse"] = resourse.text
        print(Dict_resourse)
        resourse.click()
        sleep(2)
        test_extract_data()


def test_extract_data():
    data_count = Intantiate.driver.find_elements_by_xpath("//div[@class='MuiCardContent-root lead-container']")
    print(len(data_count))
    for data in data_count:
        data.find_element_by_xpath("//div[@class='MuiCardContent-root lead-container']")
        Dict_data["data"] = data.text
        print(Dict_data)
    Intantiate.driver.find_element_by_xpath("//div[@class='bookmarked' and contains(text(), 'MissionHumane.org')]").click()
    sleep(2)

# closing browser
def test_close():
    Intantiate.driver.quit()