import openpyxl


def test_loadExcel():
    Dict = {}
    book = openpyxl.load_workbook("C:\Training matrial\Project Covid\MIssionHumanData.xlsx")
    sheet = book.active
    cell = sheet.cell(row=2, column=1)
    print(cell.value)
    for i in range(1, sheet.max_row + 1):  # to get rows
        if sheet.cell(row=i, column=2).value == "Andhra Pradesh":
            for j in range(1, sheet.max_column + 1):  # to get column
                # print(sheet.cell(row=i, column=1).value)
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                print(Dict)
